# app.py
import streamlit as st
import openpyxl
import datetime as dt
from io import BytesIO
from zipfile import ZipFile, ZIP_DEFLATED
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.workbook.properties import CalcProperties
import re
from pathlib import Path

st.set_page_config(page_title="POP files", layout="centered")
st.title("Création des fichiers POP")
st.subheader("📂 Déposez tous les fichiers POP ici")

st.markdown("""
L’application :
- 🔄 met à jour les **jours équivalents**
- 🧮 applique les **formules Excel**
- 📦 génère une **archive ZIP** contenant tous les fichiers créés
""")

# -------------------------
# Uploads (local bundled files)
# -------------------------
MAPPING_PATH = Path("Calendrier_comparatif_2026_vs_2025.xlsx")
mapping_file = BytesIO(MAPPING_PATH.read_bytes())
mapping_file.name = MAPPING_PATH.name

SOURCE_PATH = Path("Book2.xlsx")  # source template: E..AE days, AF Total formulas
source_file = BytesIO(SOURCE_PATH.read_bytes())
source_file.name = SOURCE_PATH.name

target_files = st.file_uploader("Deposer ici les fichers POP", type=["xlsx"], accept_multiple_files=True)

if target_files:
    st.info(f"📦 {len(target_files)} file(s) uploaded.")

# -------------------------
# Config
# -------------------------
HEADER_ROW_2026 = 31
HEADER_ROW_2025 = 3

START_COL = 5          # E
MAX_DAY_COL = 31       # AE (max search range for dates in row 31)
SOURCE_TOTAL_COL = 32  # AF (Total column in source template)

# Copy blocks: copy formulas for days only (E..last_day_col)
COPY_BLOCKS = [
    (32, 48),
    (58, 72),
]

# Total column is only relevant until row 48 (per your rule)
TOTAL_ROW_FROM = 32
TOTAL_ROW_TO   = 48


# -------------------------
# Helpers
# -------------------------
def copy_block(ws_src, ws_dst, r1, r2, c1, c2):
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            dst_cell = ws_dst.cell(row=row, column=col)
            dst_cell.value = ws_src.cell(row=row, column=col).value
            dst_cell.number_format = "General"  # avoid formulas-as-text


def cell_to_date(cell, default_year=None):
    v = cell.value
    if v is None:
        return None

    if cell.is_date:
        if isinstance(v, dt.datetime):
            return v.date()
        if isinstance(v, dt.date):
            return v

    if isinstance(v, (int, float)):
        try:
            d = from_excel(v)
            return d.date() if isinstance(d, dt.datetime) else d
        except Exception:
            return None

    if isinstance(v, str):
        s = v.strip().replace("'", "")
        try:
            return dt.datetime.strptime(s, "%d/%m/%Y").date()
        except ValueError:
            pass
        try:
            dm = dt.datetime.strptime(s, "%d/%m")
            if default_year is None:
                return None
            return dt.date(default_year, dm.month, dm.day)
        except ValueError:
            return None

    return None


def norm_full_date(v):
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, (int, float)):
        try:
            d = from_excel(v)
            return d.date() if isinstance(d, dt.datetime) else d
        except Exception:
            return None
    if isinstance(v, str):
        s = v.strip().replace("'", "")
        try:
            return dt.datetime.strptime(s, "%d/%m/%Y").date()
        except ValueError:
            return None
    return None


def read_mapping(file_bytes: bytes):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    mapping = {}
    for r in range(1, ws.max_row + 1):
        d26 = norm_full_date(ws.cell(row=r, column=1).value)  # A
        d25 = norm_full_date(ws.cell(row=r, column=2).value)  # B
        if d26 and d25:
            mapping[d26] = d25
    return mapping


def find_last_day_col(ws):
    """
    Find the last column (E..AE) that contains a readable date in row 31.
    """
    last = None
    for c in range(START_COL, MAX_DAY_COL + 1):
        d = cell_to_date(ws.cell(row=HEADER_ROW_2026, column=c), default_year=2026)
        if d:
            last = c
    if last is None:
        raise ValueError(f"No 2026 date found in row {HEADER_ROW_2026} between E and AE.")
    return last


def build_date_to_col(ws, header_row, end_col, default_year=None):
    result = {}
    for c in range(START_COL, end_col + 1):
        d = cell_to_date(ws.cell(row=header_row, column=c), default_year=default_year)
        if d:
            result[d] = c
    return result


def apply_mapping_formulas(ws, mapping, end_day_col):
    """
    Apply mapping formulas only on day columns E..end_day_col (NOT Total col).
    """
    col_2026 = build_date_to_col(ws, HEADER_ROW_2026, end_day_col, default_year=2026)
    col_2025 = build_date_to_col(ws, HEADER_ROW_2025, end_day_col, default_year=2025)

    for d26, d25 in mapping.items():
        if d26 not in col_2026 or d25 not in col_2025:
            continue

        tgt_col = col_2026[d26]
        src_col = col_2025[d25]

        T = get_column_letter(tgt_col)
        S = get_column_letter(src_col)

        # IMPORTANT: commas in stored formulas
        ws[f"{T}35"].value = (
            f'=IF({S}6*(1+$E$52)+{S}7*(1+$H$52)=0,0,'
            f'{S}6*(1+$E$52)+{S}7*(1+$H$52))'
        )
        ws[f"{T}41"].value = f'=IF({S}6*(1+$E$52)="","",{S}6*(1+$E$52))'
        ws[f"{T}47"].value = (
            f'=IF({S}23*(1+$N$52)+{S}24*(1+$K$52)=0,0,'
            f'{S}23*(1+$N$52)+{S}24*(1+$K$52))'
        )

def find_header_column(ws, header_name, header_row=2, start_col=1, max_col=200):
    """
    Find the column index where row `header_row` equals header_name (exact match, stripped).
    """
    for c in range(start_col, min(max_col, ws.max_column) + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        if str(v).strip() == header_name:
            return c
    raise ValueError(f"Header '{header_name}' not found in row {header_row}")

def apply_v_prod_quai_formula(ws, start_col, end_col):
    """
    Row 44 formula: =$<V_PROD_QUAI_COL>$3
    Applied from start_col to end_col.
    """
    v_prod_col = find_header_column(ws, "V_PROD_QUAI", header_row=2)
    v_prod_letter = get_column_letter(v_prod_col)

    formula = f"=${v_prod_letter}$3"

    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=44, column=c)
        cell.value = formula
        cell.number_format = "General"

def apply_prod_cam_total_formula(ws, start_col, end_col):
    """
    Row 32 formula: =$<PROD_CAM_TOTAL_COL>$3
    Applied from start_col to end_col.
    """
    prod_cam_col = find_header_column(ws, "PROD CAM TOTAL", header_row=2)
    prod_cam_letter = get_column_letter(prod_cam_col)

    formula = f"=${prod_cam_letter}$3"

    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=32, column=c)
        cell.value = formula
        cell.number_format = "General"

def apply_V_PROD_CAM_DISTRI_formula(ws, start_col, end_col):
    """
    Row 38 formula: =$<PROD_CAM_TOTAL_COL>$3
    Applied from start_col to end_col.
    """
    prod_cam_col = find_header_column(ws, "V_PROD_CAM_DISTRI", header_row=2)
    prod_cam_letter = get_column_letter(prod_cam_col)

    formula = f"=${prod_cam_letter}$3"

    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=38, column=c)
        cell.value = formula
        cell.number_format = "General"



# -------- TOTAL column rewriting (the key fix) --------

def col_letter_to_num(col: str) -> int:
    col = col.upper()
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n

def num_to_col_letter(n: int) -> str:
    return get_column_letter(n)

def replace_range_end_columns(formula: str, old_end_letters: set[str], new_end_letter: str) -> str:
    """
    Replace range ends like  ...:AE32  or ...:$AE$32  with ...:<new_end_letter>32
    Only when the END column is one of old_end_letters.
    Keeps $ and row number as-is.
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    # pattern for a range end: :$?COL$?ROW
    # group(1)= ":" , group(2)= optional "$", group(3)=COL, group(4)= optional "$", group(5)=ROW
    pat = re.compile(r'(:)(\$?)([A-Z]{1,3})(\$?)(\d+)')

    def repl(m):
        colon, dol1, col, dol2, row = m.group(1), m.group(2), m.group(3), m.group(4), m.group(5)
        if col in old_end_letters:
            return f"{colon}{dol1}{new_end_letter}{dol2}{row}"
        return m.group(0)

    return pat.sub(repl, formula)

def replace_total_self_refs(formula: str, src_total_letter: str, tgt_total_letter: str) -> str:
    """
    If the Total formula refers to the Total column letter (AF) anywhere,
    replace that column letter with the target total column letter.
    Handles AF32, $AF$32, etc.
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula
    pat = re.compile(rf'(\$?){src_total_letter}(\$?\d+)')
    return pat.sub(rf'\1{tgt_total_letter}\2', formula)

def copy_total_column_from_source(ws_src, ws_dst, target_total_col, last_day_col):
    """
    Copy Total column values/formulas from source AF into target_total_col (rows 32..48),
    while rewriting the formulas so their aggregation ends at last_day_col (not AE / not AF).
    """
    src_total_letter = get_column_letter(SOURCE_TOTAL_COL)  # "AF"
    tgt_total_letter = get_column_letter(target_total_col)  # e.g. "Y" / "AA" / ...

    last_day_letter = get_column_letter(last_day_col)
    max_day_letter  = get_column_letter(MAX_DAY_COL)        # "AE"

    # In your template, range ends might be AE (max day) or AF (if template used it by mistake)
    old_end_letters = {max_day_letter, src_total_letter}

    for r in range(TOTAL_ROW_FROM, TOTAL_ROW_TO + 1):
        src_cell = ws_src.cell(row=r, column=SOURCE_TOTAL_COL)
        dst_cell = ws_dst.cell(row=r, column=target_total_col)

        v = src_cell.value

        # rewrite formulas
        if isinstance(v, str) and v.startswith("="):
            # 1) ensure any end-of-range column becomes last_day_letter
            v = replace_range_end_columns(v, old_end_letters=old_end_letters, new_end_letter=last_day_letter)
            # 2) if formula references AF as a column (self refs), move to target total col letter
            v = replace_total_self_refs(v, src_total_letter=src_total_letter, tgt_total_letter=tgt_total_letter)

        dst_cell.value = v
        dst_cell.number_format = "General"

def clear_after_last_day(ws, last_day_col, start_row=75):
    """
    Clear all cells AFTER last_day_col, starting from start_row.
    """
    max_row = ws.max_row
    max_col = ws.max_column

    for r in range(start_row, max_row + 1):
        for c in range(last_day_col + 1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.number_format = "General"


def force_recalc_on_open(wb: openpyxl.Workbook):
    if wb.calculation is None:
        wb.calculation = CalcProperties(fullCalcOnLoad=True)
    else:
        wb.calculation.fullCalcOnLoad = True


# -------------------------
# Main action
# -------------------------
ready = target_files and len(target_files) > 0
if not ready:
    st.caption("Upload at least 1 pop file.")
    st.stop()

if st.button("✅ Apply and generate ZIP"):
    n_files = len(target_files)

    with st.spinner(f"Processing {n_files} file(s)..."):
        try:
            mapping = read_mapping(mapping_file.getvalue())
            if not mapping:
                st.error("Mapping file is empty or dates are not readable. Column A/B must contain valid dates.")
                st.stop()

            wb_src = openpyxl.load_workbook(BytesIO(source_file.getvalue()), data_only=False)
            ws_src = wb_src[wb_src.sheetnames[0]]

            zip_buffer = BytesIO()
            updated_count = 0
            total = len(target_files)

            progress = st.progress(0)

            with ZipFile(zip_buffer, "w", compression=ZIP_DEFLATED) as zf:
                for up in target_files:
                    wb_dst = openpyxl.load_workbook(BytesIO(up.getvalue()), data_only=False)
                    ws_dst = wb_dst[wb_dst.sheetnames[0]]

                    # 1) Detect last day column in target (E..AE)
                    last_day_col = find_last_day_col(ws_dst)

                    # 2) Copy day blocks from source template up to last_day_col
                    for (r1, r2) in COPY_BLOCKS:
                        copy_block(ws_src, ws_dst, r1, r2, START_COL, last_day_col)

                    # 3) Total column is right after the last day column
                    target_total_col = last_day_col + 1

                    # 4) Copy Total column from source AF -> target_total_col, but rewrite ranges to end at last_day_col
                    copy_total_column_from_source(ws_src, ws_dst, target_total_col, last_day_col)

                    # 5) Apply mapping formulas only on day columns E..last_day_col
                    apply_mapping_formulas(ws_dst, mapping, end_day_col=last_day_col)

                    # After apply_mapping_formulas(...)
                    apply_v_prod_quai_formula(ws_dst, START_COL, last_day_col)
                    apply_prod_cam_total_formula(ws_dst, START_COL, last_day_col)
                    apply_V_PROD_CAM_DISTRI_formula(ws_dst, START_COL, last_day_col)

                    # Clean everything after last day column from row 75
                    clear_after_last_day(ws_dst, last_day_col, start_row=75)

                    # 6) Force recalc
                    force_recalc_on_open(wb_dst)

                    # Save
                    out = BytesIO()
                    wb_dst.save(out)
                    file_bytes = out.getvalue()

                    original_name = up.name.rsplit(".", 1)[0]
                    zip_name = f"{original_name}_updated.xlsx"
                    zf.writestr(zip_name, file_bytes)

                    updated_count += 1
                    progress.progress(updated_count / total)

            zip_buffer.seek(0)

        except Exception as e:
            st.error(f"Error: {e}")
            st.stop()

    st.success(f"✅ All done — {updated_count}/{n_files} file(s) updated.")

    st.download_button(
        "⬇️ Download ZIP (all updated files)",
        data=zip_buffer.getvalue(),
        file_name="updated_excels.zip",
        mime="application/zip",
    )
